# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl"]
# ///
"""
サンプルExcelファイル生成スクリプト
Azure リソースパラメーター管理表（VNet + NSG + VM）を生成する。
ラボ9（Excel → IaC）で使用。
"""

import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HEADER_FONT = Font(name="Yu Gothic UI", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
BODY_FONT = Font(name="Yu Gothic UI", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def style_body(ws, start_row, end_row, col_count):
    for row in range(start_row, end_row + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER


def create_vnet_sheet(wb):
    ws = wb.create_sheet("VNet")
    headers = ["リソース名", "リージョン", "リソースグループ", "アドレス空間"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    data = [
        ["vnet-prod-001", "japaneast", "rg-network-prod", "10.0.0.0/16"],
        ["vnet-dev-001", "japaneast", "rg-network-dev", "10.1.0.0/16"],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
    style_body(ws, 2, len(data) + 1, len(headers))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22


def create_subnet_sheet(wb):
    ws = wb.create_sheet("Subnet")
    headers = ["サブネット名", "所属VNet", "アドレスプレフィックス", "用途"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    data = [
        ["snet-web", "vnet-prod-001", "10.0.1.0/24", "Webサーバー"],
        ["snet-app", "vnet-prod-001", "10.0.2.0/24", "アプリサーバー"],
        ["snet-db", "vnet-prod-001", "10.0.3.0/24", "データベース"],
        ["snet-web", "vnet-dev-001", "10.1.1.0/24", "Webサーバー（開発）"],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
    style_body(ws, 2, len(data) + 1, len(headers))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 24


def create_nsg_sheet(wb):
    ws = wb.create_sheet("NSG")
    headers = ["NSG名", "リージョン", "リソースグループ", "関連サブネット"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    data = [
        ["nsg-web-prod", "japaneast", "rg-network-prod", "snet-web"],
        ["nsg-app-prod", "japaneast", "rg-network-prod", "snet-app"],
        ["nsg-db-prod", "japaneast", "rg-network-prod", "snet-db"],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
    style_body(ws, 2, len(data) + 1, len(headers))

    # NSGルールシート
    ws2 = wb.create_sheet("NSGルール")
    rule_headers = [
        "NSG名",
        "ルール名",
        "優先度",
        "方向",
        "アクセス",
        "プロトコル",
        "送信元",
        "送信先ポート",
        "説明",
    ]
    for i, h in enumerate(rule_headers, 1):
        ws2.cell(row=1, column=i, value=h)
    style_header(ws2, 1, len(rule_headers))

    rules = [
        [
            "nsg-web-prod",
            "Allow-HTTP",
            100,
            "Inbound",
            "Allow",
            "TCP",
            "*",
            "80",
            "HTTP許可",
        ],
        [
            "nsg-web-prod",
            "Allow-HTTPS",
            110,
            "Inbound",
            "Allow",
            "TCP",
            "*",
            "443",
            "HTTPS許可",
        ],
        [
            "nsg-web-prod",
            "Deny-All-Inbound",
            4096,
            "Inbound",
            "Deny",
            "*",
            "*",
            "*",
            "その他拒否",
        ],
        [
            "nsg-app-prod",
            "Allow-From-Web",
            100,
            "Inbound",
            "Allow",
            "TCP",
            "10.0.1.0/24",
            "8080",
            "Webサブネットから許可",
        ],
        [
            "nsg-app-prod",
            "Deny-All-Inbound",
            4096,
            "Inbound",
            "Deny",
            "*",
            "*",
            "*",
            "その他拒否",
        ],
        [
            "nsg-db-prod",
            "Allow-From-App",
            100,
            "Inbound",
            "Allow",
            "TCP",
            "10.0.2.0/24",
            "1433",
            "Appサブネットから許可",
        ],
        [
            "nsg-db-prod",
            "Deny-All-Inbound",
            4096,
            "Inbound",
            "Deny",
            "*",
            "*",
            "*",
            "その他拒否",
        ],
    ]
    for r, row_data in enumerate(rules, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)
    style_body(ws2, 2, len(rules) + 1, len(rule_headers))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 20


def create_vm_sheet(wb):
    ws = wb.create_sheet("VM")
    headers = [
        "VM名",
        "リージョン",
        "リソースグループ",
        "VMサイズ",
        "OSイメージ",
        "OSディスク種類",
        "管理者ユーザー名",
        "所属サブネット",
        "関連NSG",
        "パブリックIP",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    data = [
        [
            "vm-web-prod-001",
            "japaneast",
            "rg-compute-prod",
            "Standard_D2s_v5",
            "Ubuntu 24.04 LTS",
            "Premium_LRS",
            "azureuser",
            "snet-web",
            "nsg-web-prod",
            "あり",
        ],
        [
            "vm-web-prod-002",
            "japaneast",
            "rg-compute-prod",
            "Standard_D2s_v5",
            "Ubuntu 24.04 LTS",
            "Premium_LRS",
            "azureuser",
            "snet-web",
            "nsg-web-prod",
            "あり",
        ],
        [
            "vm-app-prod-001",
            "japaneast",
            "rg-compute-prod",
            "Standard_D4s_v5",
            "Ubuntu 24.04 LTS",
            "Premium_LRS",
            "azureuser",
            "snet-app",
            "nsg-app-prod",
            "なし",
        ],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
    style_body(ws, 2, len(data) + 1, len(headers))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22


def main():
    wb = openpyxl.Workbook()
    # デフォルトの "Sheet" を削除
    wb.remove(wb.active)

    create_vnet_sheet(wb)
    create_subnet_sheet(wb)
    create_nsg_sheet(wb)
    create_vm_sheet(wb)

    output_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(output_dir, "azure-resource-params.xlsx")
    wb.save(output_path)
    print(f"サンプルExcelファイルを生成しました: {output_path}")


if __name__ == "__main__":
    main()
